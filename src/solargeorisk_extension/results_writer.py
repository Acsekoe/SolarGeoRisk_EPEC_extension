from __future__ import annotations

import os
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .model import ModelData


def _apply_sheet_formatting(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for col_idx, cell in enumerate(ws[1], start=1):
        header = "" if cell.value is None else str(cell.value)
        width = min(max(len(header), 10), 28)
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2


def _safe_get(d: Dict, k, default=0.0) -> float:
    try:
        v = d.get(k, default)
        return float(v) if v is not None else float(default)
    except Exception:
        return float(default)


def write_results_excel(
    *,
    data: ModelData,
    state: Dict[str, Dict],
    iter_rows: List[Dict[str, object]],
    detailed_iter_rows: List[Dict[str, object]] | None = None,
    output_path: str,
    meta: Dict[str, object] | None = None,
) -> None:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    Q_offer = state.get("Q_offer", {})
    x_dem = state.get("x_dem", {})
    lam = state.get("lam", {})
    x = state.get("x", {})
    tau_imp = state.get("tau_imp", {})
    tau_exp = state.get("tau_exp", {})
    obj = state.get("obj", {})

    region_rows: List[Dict[str, object]] = []
    for r in data.regions:
        imports = sum(_safe_get(x, (exp, r), 0.0) for exp in data.regions)
        exports = sum(_safe_get(x, (r, imp), 0.0) for imp in data.regions)
        region_rows.append(
            {
                "r": r,
                "Q_offer": _safe_get(Q_offer, r, 0.0),
                "x_dem": _safe_get(x_dem, r, 0.0),
                "lam": _safe_get(lam, r, 0.0),
                "obj": _safe_get(obj, r, 0.0),
                "imports": float(imports),
                "exports": float(exports),
                "Qcap": float(data.Qcap[r]),
                "Dmax": float(data.Dmax[r]),
            }
        )

    flow_rows: List[Dict[str, object]] = []
    for exp in data.regions:
        for imp in data.regions:
            flow_rows.append(
                {
                    "exp": exp,
                    "imp": imp,
                    "x": _safe_get(x, (exp, imp), 0.0),
                    "tau_imp": _safe_get(tau_imp, (imp, exp), 0.0),
                    "tau_exp": _safe_get(tau_exp, (exp, imp), 0.0),
                    "c_ship": float(data.c_ship[(exp, imp)]),
                    "c_man": float(data.c_man[exp]),
                }
            )

    df_regions = pd.DataFrame(region_rows)
    df_flows = pd.DataFrame(flow_rows)
    df_iters = pd.DataFrame(iter_rows)
    df_meta = pd.DataFrame(list((meta or {}).items()), columns=["key", "value"])
    
    # Detailed Iterations
    df_detailed = pd.DataFrame(detailed_iter_rows) if detailed_iter_rows else pd.DataFrame()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_regions.to_excel(writer, sheet_name="regions", index=False)
        df_flows.to_excel(writer, sheet_name="flows", index=False)
        df_iters.to_excel(writer, sheet_name="iters", index=False)
        if not df_detailed.empty:
            df_detailed.to_excel(writer, sheet_name="detailed_iters", index=False)
        df_meta.to_excel(writer, sheet_name="meta", index=False)

    wb = load_workbook(output_path)
    for sheet in ["regions", "flows", "iters", "detailed_iters", "meta"]:
        if sheet in wb.sheetnames:
            _apply_sheet_formatting(wb[sheet])
    wb.save(output_path)

